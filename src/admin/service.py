from pathlib import Path
import pandas as pd
import openpyxl
from sqlalchemy.orm import Session
from sqlalchemy import text
from slugify import slugify
from sqlalchemy.orm import sessionmaker
from sqlalchemy import create_engine
from src.config import settings

from src.catalog.models import Category, Product

def clean_database(db: Session):
    print("[DB] Начинаю полную очистку и сброс ID (PostgreSQL)...")
    try:
        db.execute(text("TRUNCATE TABLE products, categories RESTART IDENTITY CASCADE;"))
        db.commit()
        print("[DB] База очищена, ID сброшены.")
    except Exception as e:
        db.rollback()
        print(f"[ERROR] {e}")

def clean_price(value):
    """Очистка цены от мусора"""
    if pd.isna(value): return None
    if isinstance(value, (int, float)): return float(value)
    s_val = str(value).strip()
    if not s_val: return None
    try:
        cleaned = s_val.replace(",", ".").replace(" ", "").replace("\xa0", "")
        return float(cleaned)
    except:
        return None

def get_unique_slug(name: str, used_slugs: set) -> str:
    """Генерация уникального slug"""
    original_slug = slugify(name)
    slug = original_slug
    counter = 1
    while slug in used_slugs:
        slug = f"{original_slug}-{counter}"
        counter += 1
    used_slugs.add(slug)
    return slug

def process_excel_file(file_path: Path, db: Session):
    clean_database(db)

    print(f"[DEBUG] Начинаю обработку файла: {file_path}")

    wb = openpyxl.load_workbook(file_path, data_only=True) 
    ws = wb.active

    start_row = 1
    # Ищем строку, где есть слово "Наименование" в колонке B (индекс 2)
    # iter_rows возвращает кортежи ячеек. min_col=2 значит читаем со 2й колонки
    for row in ws.iter_rows(min_col=2, max_col=2):
        if row[0].value and "Наименование" in str(row[0].value):
            start_row = row[0].row + 1
            print(f"[DEBUG] Начало данных найдено на строке {start_row}")
            break
    
    # 4. Стек категорий
    # Хранит объекты: {'obj': CategoryORM, 'indent': int}
    category_stack = [] 
    used_slugs = set()

    # Проходим по всем строкам, начиная с данных
    # min_col=2 (Название), max_col=3 (Цена) -> получаем кортеж (cell_name, cell_price)
    for row in ws.iter_rows(min_row=start_row, min_col=2, max_col=3):
        cell_name = row[0]
        cell_price = row[1]

        raw_name = cell_name.value
        if not raw_name: # Пропускаем пустые строки
            continue
            
        name_str = str(raw_name).strip()
        if not name_str:
            continue

        price_val = clean_price(cell_price.value)
        
        # Читаем стиль
        # indent (отступ) возвращает число 0, 1, 2... (или 0.0)
        indent = int(cell_name.alignment.indent) if cell_name.alignment and cell_name.alignment.indent else 0
        
        # bold
        is_bold = cell_name.font.b if cell_name.font else False

        # --- ЛОГИКА ОПРЕДЕЛЕНИЯ: ТОВАР ИЛИ КАТЕГОРИЯ ---
        # Категория если: Нет цены ИЛИ (Жирный шрифт И цена странная/пустая)
        # В твоем файле главное отличие - отсутствие цены у категорий.
        
        is_category = (price_val is None)

        if is_category:
            # Логика стека:
            # Если новый отступ МЕНЬШЕ или РАВЕН текущему, значит мы вышли из глубины.
            # Надо убрать из стека всех, кто глубже или равен новой категории.
            
            while category_stack:
                last_cat = category_stack[-1]
                # Если отступ текущей строки <= отступа родителя в стеке -> значит тот родитель закончился
                # Например: Было "4 грани" (indent 2), пришло "2 грани" (indent 2).
                # 2 <= 2 -> выкидываем "4 грани", чтобы "2 грани" стала дочкой "Бетона" (indent 1)
                if indent <= last_cat['indent']:
                    category_stack.pop()
                else:
                    break # Нашли реального родителя (у него indent строго меньше)

            # Родитель - это тот, кто остался последним в стеке
            parent_cat = category_stack[-1]['obj'] if category_stack else None
            
            # Создаем категорию в БД
            slug = get_unique_slug(name_str, used_slugs)
            new_cat = Category(
                name=name_str, 
                slug=slug, 
                parent_id=parent_cat.id if parent_cat else None
            )
            
            db.add(new_cat)
            db.flush() # Чтобы получить ID
            
            # Добавляем в стек
            category_stack.append({'obj': new_cat, 'indent': indent})
            
            # Лог для отладки
            p_name = parent_cat.name if parent_cat else "КОРЕНЬ"
            print(f"[CAT] {indent * '-'} {name_str} (Родитель: {p_name})")

        else:
            # Товар всегда принадлежит последней открытой категории в стеке
            
            if category_stack:
                active_cat = category_stack[-1]['obj']
                
                # Защита от товаров, которые "прыгнули" влево (ошибка верстки в Excel)
                # Если у товара отступ меньше категории, возможно, надо подняться выше? 
                # Но обычно товары просто лежат в последней папке. Оставим простую логику.
                
                prod = Product(
                    name=name_str,
                    price=price_val,
                    category_id=active_cat.id
                )
                
                db.add(prod)
            else:
                print(f"[WARN] Товар '{name_str}' без категории! Пропускаем.")

    db.commit()
    wb.close()
    print("[DEBUG] Импорт завершен.")

sync_engine = create_engine(settings.DATABASE_URL.replace("+asyncpg", ""))
SyncSession = sessionmaker(bind=sync_engine)

def handle_price_upload(file_content: bytes):
    """
    Эта функция выполняется в отдельном потоке.
    """
    try:
        print("⏳ [Background] Начало обработки прайса...")
        
        dest_path = settings.UPLOAD_DIR / "price.xlsx"
        
        dest_path.parent.mkdir(parents=True, exist_ok=True)

        with open(dest_path, "wb") as buffer:
            buffer.write(file_content)
        
        print("💾 [Background] Файл сохранен на диск.")

        with SyncSession() as db:
            process_excel_file(dest_path, db)
            
        print("✅ [Background] Прайс успешно загружен и распарсен!")

    except Exception as e:
        print(f"❌ [Background Error] Ошибка обработки прайса: {e}")